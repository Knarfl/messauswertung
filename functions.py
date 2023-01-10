import pandas as pd
import os
import plotly.express as px
import openpyxl
import drive as dv
import streamlit as st

from io import BytesIO
from openpyxl.writer.excel import save_virtual_workbook
from openpyxl.utils.dataframe import dataframe_to_rows


def upload(uploaded_files):
    list_data = list()
    list_keys = list()
    for uploaded_file in uploaded_files:
        keys, extension = os.path.splitext(uploaded_file.name)
        list_keys.append(int(keys))
        # Lese die hochgeladene Datei in ein Pandas DataFrame
        df = pd.read_excel(uploaded_file, skiprows=4, usecols='A:X', nrows=8)
        # Zeige den Inhalt des DataFrames im GUI an
        list_data.append(df.dropna(subset=["Network Provider"]))

    dict_data = dict(zip(list_keys, list_data))
    print(list_keys)
    return list_data, dict_data


def find_all_lac(list_data):
    list_lac = []
    list_car = []

    for element in list_data:
        list_lac.extend(element["TETRA LA or DMO Source Address"].to_list())
        list_car.extend(element["Main Carrier"].to_list())

    s_values = pd.Series(list_lac)
    s_values = s_values.value_counts(sort=True)

    c_value = pd.Series(list_car)
    c_value = c_value.value_counts(sort=True)

    dict_lac = dict(zip(list(s_values.index), list(c_value.index)))

    return list(s_values.index), dict_lac


def check_unique(list_dataframe):
    df_count = list_dataframe.value_counts(sort=True)
    list_to_check = []

    for element in df_count:
        if element > 1:
            list_to_check.extend(df_count.index[df_count == element].tolist())

    return list_to_check


def create_dataframe(list_lac):
    list_index = list(range(0, 360, 15))
    df = pd.DataFrame(data=[], columns=list_lac, index=list_index)
    return df


def fill_dataframe(df, list_data, type_filter, replace_nan):
    i = 0
    for dataframe in list_data:
        dict_temp = {}
        for element in dataframe["TETRA LA or DMO Source Address"]:
            list_to_check = check_unique(dataframe["TETRA LA or DMO Source Address"])

            if element in list_to_check:
                df_check = dataframe.loc[dataframe["TETRA LA or DMO Source Address"] == element, type_filter]
                max_value = max(df_check.tolist())
                dict_temp[element] = max_value

            else:
                dict_temp[element] = \
                    dataframe.loc[dataframe["TETRA LA or DMO Source Address"] == element, type_filter].values[0]

        new_df = pd.DataFrame(dict_temp, index=[i])
        df.update(new_df)
        i = i + (360 / len(list_data))

    return df.fillna(replace_nan)


def create_fin_dataframe(dict_data, type_filter, replace_nan):
    # type_filter = "TETRA Power Median Average"  # "TETRA Power Median Max"
    list_angle = list(dict_data.keys())
    dict_table = dict.fromkeys(list_angle, dict())
    list_all_lac = list()
    list_all_car = list()

    for key in list_angle:
        list_lac = dict_data[key]["TETRA LA or DMO Source Address"].to_list()
        list_power = dict_data[key][type_filter].to_list()
        list_car = dict_data[key]["Main Carrier"].to_list()

        list_all_lac.extend(list_lac)
        list_all_lac = [x for x in list_all_lac if pd.isnull(x) == False]
        list_all_car.extend(list_car)
        list_all_car = [x for x in list_all_car if pd.isnull(x) == False]

        dict_table[key] = dict(zip(list_lac, list_power))

    dic_lac = dict(zip(list_all_lac, list_all_car))

    df_data = pd.DataFrame.from_dict(dict_table, orient='index')
    df_data = df_data.dropna(axis=1, how='all')
    df_data = df_data.fillna(replace_nan)
    df_data = df_data.sort_index()

    return df_data, dic_lac


@st.cache()
def upload_files(uploaded_files):
    dict_data = dict()

    for file in uploaded_files:
        keys, extension = os.path.splitext(file.name)
        separator = "_"
        list_name = keys.split(separator)
        angle = int(list_name[-1])

        if extension == ".xls" or extension == ".xlsx":
            df = pd.read_excel(file, skiprows=4, usecols='A:X', nrows=8)
            dict_data.update({angle: df})

    return dict_data


def show_polar_chart(df):
    df_new = pd.melt(df, ignore_index=False)
    df_new = df_new.reset_index(col_fill=1)

    df_new.columns = ['Winkel', 'LAC', 'Pegel']

    fig = px.line_polar(df_new, r="Pegel", theta="Winkel", color="LAC", line_close=True, template="plotly_dark")
    fig_linear = px.line(df_new, x='Winkel', y='Pegel', color='LAC', template="plotly_dark")
    return fig, fig_linear


def create_xlsx(filename, dict_form, df, dict_lac, filename_tetra):
    file = dv.get_file(filename)

    with open(filename, "wb+") as f:
        for chunk in file.iter_chunks(4096):
            f.write(chunk)

    wb = openpyxl.load_workbook(filename)
    ws = wb.active

    company = dict_form["company"]
    title = "Panoramamessung"
    creator = dict_form["creator"]
    date = dict_form["date"]
    project = dict_form["project"]
    coordinates = dict_form["coordinates"]
    h_nn = dict_form["h_nn"]
    h_a = dict_form["h_a"]
    gain_a = dict_form["gain_a"]
    attenuation = dict_form["attenuation"]

    ws.oddHeader.left.text = f"ausführende Firma:\n{company}"
    ws.oddHeader.center.text = title
    ws.oddFooter.left.text = f"Name des Verantwortlichen: {creator}\n Datum und Uhrzeit: {date}"
    ws.oddFooter.right.text = f"{project}\n Koordinaten in WGS84: {coordinates}"

    ws.cell(row=24, column=3).value = h_nn
    ws.cell(row=25, column=3).value = h_a
    ws.cell(row=24, column=8).value = gain_a
    ws.cell(row=25, column=8).value = attenuation

    rows = dataframe_to_rows(df, index=False, header=True)
    for r_idx, row in enumerate(rows, 1):
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx + 32, column=c_idx + 2, value=value)

    ####
    file = dv.get_file(filename_tetra)

    with open(filename_tetra, "wb+") as f:
        for chunk in file.iter_chunks(4096):
            f.write(chunk)

    df_tetra = pd.read_excel("TETRA BOS Frequenztabelle.xlsx", sheet_name="Daten",
                             usecols=["Kanalnr.", "Carrier", "Frequenz \nDownlink"])
    df_tetra.columns = ['Kanal', 'Carrier', 'Frequenz']

    list_car = []
    list_lac = list(df.columns)
    for key in list_lac:
        list_car.append(dict_lac[key])

    x = 0
    for element in list_car:
        ws.cell(row=31, column=x + 3).value = element
        ws.cell(row=32, column=x + 3).value = df_tetra.loc[df_tetra.Carrier == element, 'Frequenz'].values[0]
        x = x + 1

    ######
    list_head = list(df.columns.values)
    len_head = len(list_head)
    if len_head <= 15:
        index = df.index
        for x in range(0, len_head):
            val_max = df[list_head[x]].max()
            indexes = index.get_indexer(df.loc[df[list_head[x]] == val_max].index)
            list_index = df.head(24).index.tolist()
            print(list_index)
            ws.cell(row=16, column=x+3).value = list_head[x]  # LAC
            ws.cell(row=17, column=x+3).value = list_index[indexes[0]]    # angle
            ws.cell(row=18, column=x+3).value = df[list_head[x]].max()    # max power of LAC

        processed_data = BytesIO(save_virtual_workbook(wb))
    else:
        diff = len_head - 15
        processed_data = diff

    return processed_data
